from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
import datetime
import os
import glob
from dateutil.relativedelta import relativedelta
from pathlib import Path


def convert_to_seconds(time_string):
    h, m, s = map(int, time_string.split(':'))
    return h * 3600 + m * 60 + s


def get_unique_filename(filepath: str) -> str:
    filename, extension = os.path.splitext(filepath)
    counter = 1

    while os.path.exists(filepath):
        filepath = f"{filename}({counter}){extension}"
        counter += 1

    return filepath


def read_config(file_path: Path):
    """
    Function to read the configuration file and extract constants.
    """
    constants = {}

    with open(file_path, 'r') as file:
        for line in file:
            key, value = line.strip().split('=')
            constants[key] = int(value)

    return constants


def read_data(file_path):
    df = pd.read_excel(file_path)
    return df


def split_driver_tags(df):
    df[['Location', 'Peer Group']] = df['Driver Tags'].str.split(',', expand=True)
    df[['Location', 'Peer Group']] = df[['Location', 'Peer Group']].apply(lambda x: x.str.strip())
    return df


def create_filtered_report(df, tag, columns):
    """
    Function to create a report for drivers with a specific tag.
    """
    # Filter the DataFrame based on the tag
    filtered_df = df[df['Driver Tags'].str.contains(tag, na=False)]

    # Create a dictionary for the report
    report_dict = {}

    # Select columns for the report
    for column in columns:
        report_dict[column] = filtered_df[column]

    # Create the report DataFrame
    filtered_report = pd.DataFrame(report_dict)

    return filtered_report


def export_to_excel(driver_safety_report, output_path):
    with pd.ExcelWriter(output_path) as writer:
        driver_safety_report.to_excel(writer, sheet_name='Driver Safety', index=False)


def get_latest_file_in_directory(directory, *extensions):
    # Find all files in the directory with any of the specified extensions
    files = []
    for extension in extensions:
        files.extend(glob.glob(f"{directory}/*.{extension}"))

    # Get the newest file (based on modification time)
    newest_file = max(files, key=os.path.getmtime)

    return newest_file


def score_range(score):
    if score == 100:
        return "Perfect 100"
    elif score >= 70:
        return "Above 70"
    elif 36 <= score < 70:
        return "Below 70"
    elif score <= 35:
        return "Critical - Below 35"


def main():
    """
    Main function to read the data and generate driver safety report.
    """
    # Get the directory of the current script
    directory = os.path.dirname(os.path.abspath(__file__))

    # Find all .xlsx files in the directory
    xlsx_files = glob.glob(f"{directory}/Samsara _raw_data/*.xlsx")

    if not xlsx_files:
        print(f"No .xlsx files found in directory: {directory}")
        return

    # Select the first .xlsx file found
    input_file_path = xlsx_files[0]

    # Read the config file
    config_path = 'Samsara _raw_data/config.txt'
    config = read_config(config_path)

    # Read the data
    df = read_data(input_file_path)

    # Split the driver tags
    df = split_driver_tags(df)

    # Add new column of drive times in seconds
    df['Drive Time (seconds)'] = df['Drive Time (hh:mm:ss)'].apply(convert_to_seconds)

    # Filter out drivers with less than the configured minimum drive time
    df = df[df['Drive Time (seconds)'] >= config['MIN_DRIVE_TIME']]

    # Make the summed up columns
    df['Collision Risk'] = df['Following Distance'] + df['Late Response (Manual)'] + df['Near Collision (Manual)']
    df['Harsh Events'] = df['Harsh Accel'] + df['Harsh Brake'] + df['Harsh Turn']
    df['Traffic Violations'] = df['Rolling Stop'] + df['Did Not Yield (Manual)'] + df['Ran Red Light (Manual)'] + \
                               df['Lane Departure (Manual)']
    df['Policy Violations'] = df['Obstructed Camera (Automatic)'] + df['Obstructed Camera (Manual)'] + df[
        'Eating/Drinking (Manual)'] + df['Smoking (Manual)'] + df['No Seat Belt']
    # df['Speeding %'] = df['Percent Moderate Speeding'] + df['Percent Heavy Speeding'] + df['Percent Severe Speeding']
    df['Score Range'] = df['Safety Score'].apply(score_range)
    df['Moderate Speeding'] = df['Time Over Speed Limit (hh:mm:ss) - Moderate']
    df['Heavy Speeding'] = df['Time Over Speed Limit (hh:mm:ss) - Heavy']
    df['Severe Speeding'] = df['Time Over Speed Limit (hh:mm:ss) - Severe']

    # Create a filtered DataFrame for each report
    driver_scorecard = df[df['Driver Tags'].str.contains("Driver|Reset|Warehouse", na=False)].copy()
    manager_scorecard = df[df['Driver Tags'].str.contains("Manager", na=False)].copy()

    # Define columns for the driver scorecard dataframe
    scorecard_columns = ['Score Range', 'Location', 'Driver Name', 'Peer Group',
                         'Safety Score', 'Drive Time (hh:mm:ss)', 'Moderate Speeding',
                         'Heavy Speeding', 'Severe Speeding',
                         'Mobile Usage', 'Crash', 'Collision Risk', 'Harsh Events',
                         'Inattentive Driving', 'Traffic Violations', 'Policy Violations']

    driver_scorecard = driver_scorecard.loc[:, scorecard_columns]
    manager_scorecard = manager_scorecard.loc[:, scorecard_columns]

    # Prepare the reports list with titles and dataframes
    reports = [
        {'title': 'Driver Scorecard', 'dataframe': driver_scorecard},
        {'title': 'Manager Scorecard', 'dataframe': manager_scorecard},

    ]

    # Define the workbook in memory
    wb = Workbook()

    # remove the default sheet created and keep our sheets only
    wb.remove(wb.active)

    # create reports (sheets) and add them to the workbook
    for report in reports:
        ws = wb.create_sheet(title=report['title'])

        # add data to the sheet
        for r in dataframe_to_rows(report['dataframe'], index=False, header=True):
            ws.append(r)

    # Load the existing workbook (template)
    wb = load_workbook('template/template.xlsx')

    # create reports (sheets) and add them to the workbook
    for report in reports:
        ws = wb[report['title']]  # get the sheet by name

        # clear existing data in the sheet
        for row in ws.iter_rows(min_row=14, max_row=ws.max_row):  # start from row 14 to preserve headers
            for cell in row:
                cell.value = None

        # add data to the sheet
        for r_index, r in enumerate(dataframe_to_rows(report['dataframe'], index=False, header=False),
                                    14):  # start from row 14
            for c_index, value in enumerate(r, 1):
                ws.cell(row=r_index, column=c_index, value=value)

    # Get the previous month
    current_month = datetime.datetime.now() - relativedelta(months=0)

    # Format the output file path
    directory_path = Path(directory)
    output_file_path = directory_path.parent / 'MTD Safety Scorecard/Report' / \
                       f'MTD Safety Scorecard - {current_month.strftime("%d %b %Y")}.xlsx'

    # If file already exists, append a number suffix
    directory_path = Path(directory)
    if output_file_path.is_file():
        counter = 1
        while output_file_path.is_file():
            output_file_path = (directory_path.parent / 'MTD Safety Scorecard/Report' /
                                f' MTD Safety Scorecard - {current_month.strftime("%d %b %Y")} ({counter}).xlsx')
            counter += 1

    # Save the workbook
    wb.save(output_file_path)


if __name__ == "__main__":
    main()
